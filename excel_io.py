# -*- coding: utf-8 -*-
"""Excel export/import for Dream Academy Manager (openpyxl, Arabic-safe)."""
import csv
import io
import os
from datetime import date, datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _sheet(wb, title, headers, rows, rtl=True):
    ws = wb.create_sheet(title)
    ws.sheet_view.rightToLeft = rtl
    ws.append(headers)
    for c in ws[1]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append(list(r))
    # auto column widths
    for i, h in enumerate(headers, 1):
        width = len(str(h)) + 4
        for r in rows[:200]:
            v = r[i - 1] if i - 1 < len(r) else ""
            width = max(width, min(len(str(v)) + 4, 45))
        ws.column_dimensions[get_column_letter(i)].width = width
    return ws


def export_all(con):
    """Return xlsx bytes with Players / Subscriptions / Payments / Attendance / Summary sheets."""
    wb = Workbook()
    wb.remove(wb.active)

    players = con.execute(
        "SELECT p.id, p.full_name, p.birth_date, p.gender, p.phone, p.guardian_name, p.guardian_phone, "
        "g.name_ar AS grp, p.join_date, p.status, p.notes FROM players p LEFT JOIN groups g ON g.id=p.group_id "
        "ORDER BY p.full_name"
    ).fetchall()
    _sheet(wb, "Players اللاعبين",
           ["ID", "الاسم", "تاريخ الميلاد", "الجنس", "الهاتف", "ولي الأمر", "هاتف ولي الأمر",
            "المجموعة", "تاريخ الانضمام", "الحالة", "ملاحظات"],
           [tuple(r) for r in players])

    subs = con.execute(
        "SELECT s.id, p.full_name, s.start_date, s.sessions_total, s.sessions_used, s.price, s.expiry_date, s.status "
        "FROM subscriptions s JOIN players p ON p.id=s.player_id ORDER BY s.start_date DESC"
    ).fetchall()
    _sheet(wb, "Subscriptions الاشتراكات",
           ["ID", "اللاعب", "تاريخ البداية", "الحصص الكلية", "الحصص المستخدمة", "السعر", "تاريخ الانتهاء", "الحالة"],
           [tuple(r) for r in subs])

    pays = con.execute(
        "SELECT pm.receipt_no, p.full_name, pm.amount, pm.date, pm.method, pm.note "
        "FROM payments pm JOIN players p ON p.id=pm.player_id ORDER BY pm.date DESC"
    ).fetchall()
    _sheet(wb, "Payments الدفعات",
           ["رقم الإيصال", "اللاعب", "المبلغ (دينار)", "التاريخ", "طريقة الدفع", "ملاحظة"],
           [tuple(r) for r in pays])

    att = con.execute(
        "SELECT a.session_date, p.full_name, g.name_ar, a.status, a.unpaid, a.marked_by "
        "FROM attendance a JOIN players p ON p.id=a.player_id LEFT JOIN groups g ON g.id=a.group_id "
        "ORDER BY a.session_date DESC"
    ).fetchall()
    _sheet(wb, "Attendance الحضور",
           ["التاريخ", "اللاعب", "المجموعة", "الحالة", "غير مدفوع", "سجّله"],
           [(r[0], r[1], r[2], {"present": "حاضر", "absent": "غائب", "excused": "معذور"}.get(r[3], r[3]),
             "نعم" if r[4] else "", r[5]) for r in att])

    # expenses — where the money went
    try:
        exp = con.execute(
            "SELECT date, COALESCE(NULLIF(category,''),'other') c, amount, note FROM expenses "
            "ORDER BY date DESC").fetchall()
    except Exception:
        exp = []
    _sheet(wb, "Expenses المصاريف",
           ["التاريخ / Date", "البند / Category", "المبلغ / Amount (JD)", "الوصف / Description"],
           [tuple(r) for r in exp])

    # coaches & salaries
    try:
        coaches = con.execute(
            "SELECT name, phone, salary_type, salary_amount, active FROM coaches ORDER BY name").fetchall()
    except Exception:
        coaches = []
    _sheet(wb, "Coaches المدربين",
           ["الاسم / Name", "الهاتف / Phone", "نوع الراتب / Salary type", "الراتب / Amount", "فعال / Active"],
           [(r[0], r[1], r[2], r[3], "نعم" if r[4] else "لا") for r in coaches])

    month = date.today().strftime("%Y-%m")
    revenue = con.execute("SELECT COALESCE(SUM(amount),0) s FROM payments WHERE date LIKE ?",
                          (month + "%",)).fetchone()["s"]
    try:
        other_exp = con.execute("SELECT COALESCE(SUM(amount),0) s FROM expenses WHERE date LIKE ?",
                                (month + "%",)).fetchone()["s"]
    except Exception:
        other_exp = 0
    kpis = [
        ("اللاعبين الفعالين / Active players",
         con.execute("SELECT COUNT(*) c FROM players WHERE status='active'").fetchone()["c"]),
        ("إيراد هذا الشهر / Revenue this month (JD)", revenue),
        ("مصاريف هذا الشهر / Expenses this month (JD)", other_exp),
        ("اشتراكات فعالة / Active subscriptions",
         con.execute("SELECT COUNT(*) c FROM subscriptions WHERE status='active'").fetchone()["c"]),
        ("حصص غير مدفوعة / Unpaid sessions",
         con.execute("SELECT COUNT(*) c FROM attendance WHERE unpaid=1").fetchone()["c"]),
        ("تاريخ التصدير / Exported at", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    _sheet(wb, "Summary الملخص", ["البند / Item", "القيمة / Value"], kpis)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


PLAYER_TEMPLATE_HEADERS = ["الاسم الكامل*", "تاريخ الميلاد (YYYY-MM-DD)", "الجنس (M/F)", "هاتف اللاعب",
                           "اسم ولي الأمر", "هاتف ولي الأمر (07XXXXXXXX)", "المجموعة (اسمها بالعربي)", "ملاحظات"]


def players_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Players"
    ws.sheet_view.rightToLeft = True
    ws.append(PLAYER_TEMPLATE_HEADERS)
    for c in ws[1]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
    ws.append(["عمر خالد", "2013-05-01", "M", "", "خالد أحمد", "0791234567", "ناشئين", "مثال — احذف هذا الصف"])
    for i in range(1, len(PLAYER_TEMPLATE_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 26
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _norm_cell(v):
    return str(v).strip() if v is not None else ""


def _rows_from_xlsx(file_stream):
    wb = load_workbook(file_stream)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and any(c is not None and str(c).strip() for c in row):
            yield [_norm_cell(c) for c in row]


def _rows_from_delimited(file_stream, delimiter):
    raw = file_stream.read()
    if isinstance(raw, bytes):
        # handle a UTF-8 BOM from Excel-exported CSVs
        raw = raw.decode("utf-8-sig", errors="replace")
    lines = [ln for ln in raw.splitlines() if ln.strip()]
    if not lines:
        return
    reader = csv.reader(lines, delimiter=delimiter)
    rows = list(reader)
    # drop a header row if the first cell looks like a header label
    head = (rows[0][0] if rows and rows[0] else "").strip().lower()
    if head in ("name", "full name", "full_name", "الاسم", "الاسم الكامل", "الاسم الكامل*"):
        rows = rows[1:]
    for r in rows:
        if r and r[0].strip():
            yield [c.strip() for c in r]


def _rows_from_txt(file_stream):
    raw = file_stream.read()
    if isinstance(raw, bytes):
        raw = raw.decode("utf-8-sig", errors="replace")
    for ln in raw.splitlines():
        name = ln.strip()
        if name:
            yield [name]


def import_players(con, file_storage, suggest_group_fn, filename=None):
    """Bulk-load players from .xlsx / .csv / .tsv / .txt.

    Column order (when present): name, birth_date, gender, phone, guardian_name,
    guardian_phone, group, notes. TXT and single-column files import names only.
    Returns (imported, skipped, errors).
    """
    filename = (filename or getattr(file_storage, "filename", "") or "").lower()
    stream = getattr(file_storage, "stream", file_storage)

    # sniff the leading bytes so a raw BytesIO (no filename) is still detected
    magic = b""
    try:
        pos = stream.tell()
        head = stream.read(4)
        stream.seek(pos)
        magic = head if isinstance(head, bytes) else head.encode("utf-8", "replace")
    except Exception:
        pass
    is_xlsx = filename.endswith(".xlsx") or magic[:2] == b"PK"

    if is_xlsx:
        row_iter = _rows_from_xlsx(stream)
    elif filename.endswith(".tsv"):
        row_iter = _rows_from_delimited(stream, "\t")
    elif filename.endswith(".txt"):
        row_iter = _rows_from_txt(stream)
    else:
        # csv (default for anything text-like, incl. unknown/no extension)
        row_iter = _rows_from_delimited(stream, ",")

    imported, skipped, errors = 0, 0, []
    groups = {}
    for r in con.execute("SELECT id, name_ar, name_en FROM groups").fetchall():
        groups[(r["name_ar"] or "").strip()] = r["id"]
        groups[(r["name_en"] or "").strip().lower()] = r["id"]

    for i, row in enumerate(row_iter, start=2):
        name = (row[0] if row else "").strip()
        if not name or "احذف هذا الصف" in (row[7] if len(row) > 7 else ""):
            continue
        try:
            g = lambda idx: (row[idx].strip() if len(row) > idx and row[idx] else "")
            birth = g(1)[:10]
            gd = g(2).upper()
            gender = "F" if gd.startswith("F") or "أنثى" in gd or gd == "F" else "M"
            phone = g(3)
            gname = g(4)
            gphone = g(5)
            grp_name = g(6)
            notes = g(7)
            group_id = groups.get(grp_name) or groups.get(grp_name.lower()) or suggest_group_fn(con, birth, gender)
            if con.execute("SELECT id FROM players WHERE full_name=?", (name,)).fetchone():
                skipped += 1
                continue
            con.execute(
                "INSERT INTO players (full_name,birth_date,gender,phone,guardian_name,guardian_phone,"
                "group_id,join_date,notes,status) VALUES (?,?,?,?,?,?,?,?,?,'active')",
                (name, birth, gender, phone, gname, gphone, group_id, date.today().isoformat(), notes),
            )
            imported += 1
        except Exception as e:
            errors.append(f"row {i}: {e}")
    con.commit()
    return imported, skipped, errors


def weekly_auto_export(con, exports_dir):
    """Write a dated full export to /exports if none exists in the last 7 days."""
    os.makedirs(exports_dir, exist_ok=True)
    existing = sorted(f for f in os.listdir(exports_dir) if f.startswith("academy-export-"))
    if existing:
        last = existing[-1][len("academy-export-"):len("academy-export-") + 10]
        try:
            if (date.today() - date.fromisoformat(last)).days < 7:
                return None
        except ValueError:
            pass
    path = os.path.join(exports_dir, f"academy-export-{date.today().isoformat()}.xlsx")
    with open(path, "wb") as f:
        f.write(export_all(con))
    return path
